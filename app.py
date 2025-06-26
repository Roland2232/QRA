from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import qrcode
import numpy as np
import json
import os
import secrets
import string
import socket
import subprocess
from datetime import datetime, timedelta
from PIL import Image
from geopy.distance import geodesic
import pymysql
from models import db, Admin, Teacher, Course, Student, AttendanceSession, Attendance
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from functools import wraps

def get_network_ip():
    """Get the actual network IP address that mobile devices can connect to"""
    try:
        
        if os.environ.get('FLASK_ENV') == 'development':
            return 'localhost'
            
        
        result = subprocess.run(['ipconfig'], capture_output=True, text=True, shell=True)
        lines = result.stdout.split('\n')
        
        for i, line in enumerate(lines):
            if 'Wireless LAN adapter Wi-Fi:' in line or 'Wi-Fi:' in line:
                
                for j in range(i, min(i + 10, len(lines))):
                    if 'IPv4 Address' in lines[j] and ':' in lines[j]:
                        ip = lines[j].split(':')[1].strip()
                        if ip and not ip.startswith('127.') and '.' in ip:
                            return ip
        
        # Fallback method
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        return local_ip
    except Exception as e:
        print(f"Error getting network IP: {e}")
        return 'localhost'  # Default to localhost if all else fails

# Get the actual network IP automatically
NETWORK_IP = get_network_ip()
print(f"🌐 Detected Network IP: {NETWORK_IP}")

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-super-secret-key-change-this-in-production'

# Server Configuration - Use detected IP for network access
app.config['SERVER_NAME'] = None  # Allow flexible host names
app.config['SERVER_IP'] = NETWORK_IP  # Use detected network IP address
app.config['SERVER_PORT'] = 5000

print(f"🌐 Server will be accessible at: http://{app.config['SERVER_IP']}:{app.config['SERVER_PORT']}")
print(f"📱 QR codes will point to this address for mobile access")
print(f"🔗 Other devices can access via: http://{app.config['SERVER_IP']}:{app.config['SERVER_PORT']}")
print(f"💻 Local access still available at: http://localhost:{app.config['SERVER_PORT']}")

# Database Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://remi:1234@localhost:3307/Qra'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Email Configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'notorios2003@gmail.com'
app.config['MAIL_PASSWORD'] = 'thsl usar tiol uvxi'
app.config['MAIL_DEFAULT_SENDER'] = 'notorios2003@gmail.com'

# File Upload Configuration
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
mail = Mail(app)

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/qr_codes', exist_ok=True)

def teacher_password_required(f):
    """Decorator to ensure teacher has changed their initial password"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if isinstance(current_user, Teacher) and current_user.must_change_password:
            flash('You must change your password before accessing this feature', 'warning')
            return redirect(url_for('change_password'))
        return f(*args, **kwargs)
    return decorated_function

@login_manager.user_loader
def load_user(user_id):
    user = Admin.query.get(user_id)
    if user:
        return user
    return Teacher.query.get(user_id)

# Utility Functions
def generate_password(length=12):
    characters = string.ascii_letters + string.digits + "!@#$%^&*"
    return ''.join(secrets.choice(characters) for _ in range(length))

def send_email(to, subject, template, **kwargs):
    msg = Message(
        subject=subject,
        recipients=[to],
        html=template,
        sender=app.config['MAIL_DEFAULT_SENDER']
    )
    try:
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

def generate_qr_code(data, filename):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    filepath = os.path.join('static/qr_codes', filename)
    img.save(filepath)
    return filepath

def calculate_distance(lat1, lon1, lat2, lon2):
    return geodesic((lat1, lon1), (lat2, lon2)).meters

def get_external_url(endpoint, **values):
    """Generate external URL with the server's IP address"""
    with app.test_request_context():
        path = url_for(endpoint, **values)
    return f"http://{app.config['SERVER_IP']}:{app.config['SERVER_PORT']}{path}"

# Make external URL function available in templates
@app.template_global()
def external_url(endpoint, **values):
    """Template global function for generating external URLs"""
    return get_external_url(endpoint, **values)

# Make server config available in templates
@app.context_processor
def inject_server_config():
    return {
        'SERVER_IP': app.config['SERVER_IP'],
        'SERVER_PORT': app.config['SERVER_PORT']
    }

# Routes
@app.route('/')
def index():
    # Always logout any existing session and redirect to login
    logout_user()
    return redirect(url_for('login'))

@app.route('/home')
def home():
    """Original home page moved to /home route"""
    return render_template('index.html')

@app.route('/mobile-test')
def mobile_test():
    """Simple test page to verify mobile access"""
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Mobile Test - QR Attendance</title>
        <style>
            body {{ 
                font-family: Arial, sans-serif; 
                padding: 20px; 
                text-align: center; 
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                min-height: 100vh;
                margin: 0;
            }}
            .container {{
                max-width: 400px;
                margin: 0 auto;
                background: white;
                color: #333;
                padding: 30px;
                border-radius: 15px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.3);
            }}
            .success {{ 
                color: #28a745; 
                font-size: 24px; 
                margin: 20px 0; 
                font-weight: bold;
            }}
            .info {{ 
                background: #f8f9fa; 
                padding: 20px; 
                border-radius: 10px; 
                margin: 20px 0; 
                border-left: 4px solid #007bff;
            }}
            .btn {{
                background: #007bff;
                color: white;
                padding: 12px 24px;
                border: none;
                border-radius: 25px;
                text-decoration: none;
                display: inline-block;
                margin: 10px;
                font-size: 16px;
                transition: all 0.3s;
            }}
            .btn:hover {{
                background: #0056b3;
                transform: translateY(-2px);
            }}
            h1 {{ margin-top: 0; }}
            @media (max-width: 480px) {{
                body {{ padding: 10px; }}
                .container {{ padding: 20px; }}
                .success {{ font-size: 20px; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📱 Mobile Access Test</h1>
            <div class="success">✅ SUCCESS!</div>
            <p>Your mobile device can access this server!</p>
            
            <div class="info">
                <h3>🌐 Connection Details</h3>
                <p><strong>Server:</strong> {app.config['SERVER_IP']}:{app.config['SERVER_PORT']}</p>
                <p><strong>Status:</strong> ✅ Connected</p>
            </div>
            
            <div class="info">
                <h3>📋 Next Steps</h3>
                <p>1. Login with admin code: <strong>23456</strong></p>
                <p>2. Create teacher accounts</p>
                <p>3. Generate QR codes for courses</p>
                <p>4. Students can scan & register from their phones!</p>
            </div>
            
            <a href="/" class="btn">🚀 Go to Login</a>
            <a href="/home" class="btn">🏠 View Home</a>
        </div>
    </body>
    </html>
    """

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Check if it's admin login (has admin_code field)
        if 'admin_code' in request.form:
            admin_code = request.form['admin_code']
            
            if admin_code == '23456':
                # Create a temporary admin session
                admin = Admin.query.first()
                if not admin:
                    # Create admin if doesn't exist
                    admin = Admin(
                        username='admin',
                        email='admin@qrattendance.com',
                        password_hash=generate_password_hash('admin123')
                    )
                    db.session.add(admin)
                    db.session.commit()
                
                login_user(admin)
                flash('Admin access granted')
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid admin code')
                return redirect(url_for('login'))
                
        # Check if it's teacher login (has username and password fields)
        elif 'username' in request.form and 'password' in request.form:
            username = request.form['username']
            password = request.form['password']
            
            teacher = Teacher.query.filter_by(username=username).first()
            
            if teacher and check_password_hash(teacher.password_hash, password):
                login_user(teacher)
                
                # Check if teacher must change password on first login
                if teacher.must_change_password:
                    flash('You must change your password before continuing', 'warning')
                    return redirect(url_for('change_password'))
                
                flash('Welcome back!')
                return redirect(url_for('teacher_dashboard'))
            else:
                flash('Invalid credentials')
                return redirect(url_for('login'))
        else:
            flash('Invalid login attempt')
            return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/admin_login', methods=['POST'])
def admin_login():
    return redirect(url_for('login'))

@app.route('/teacher_login', methods=['POST'])
def teacher_login():
    username = request.form['username']
    password = request.form['password']
    
    teacher = Teacher.query.filter_by(username=username).first()
    
    if teacher and check_password_hash(teacher.password_hash, password):
        login_user(teacher)
        
        # Check if teacher must change password
        if teacher.must_change_password:
            flash('You must change your password before continuing', 'warning')
            return redirect(url_for('change_password'))
        
        return redirect(url_for('teacher_dashboard'))
    else:
        flash('Invalid credentials')
        return redirect(url_for('login'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if not isinstance(current_user, Admin):
        flash('Access denied')
        return redirect(url_for('index'))
    
    teachers = Teacher.query.all()
    return render_template('admin_dashboard.html', teachers=teachers)

@app.route('/admin/create_teacher', methods=['GET', 'POST'])
@login_required
def create_teacher():
    if not isinstance(current_user, Admin):
        flash('Access denied')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        full_name = request.form['full_name']
        
        if Teacher.query.filter_by(username=username).first():
            flash('Username already exists')
            return render_template('create_teacher.html')
        
        if Teacher.query.filter_by(email=email).first():
            flash('Email already exists')
            return render_template('create_teacher.html')
        
        password = generate_password()
        password_hash = generate_password_hash(password)
        
        teacher = Teacher(
            username=username,
            email=email,
            password_hash=password_hash,
            full_name=full_name,
            created_by=current_user.id
        )
        
        db.session.add(teacher)
        db.session.commit()
        
        email_template = f"""
        <h2>Your Teacher Account Has Been Created</h2>
        <p>Dear {full_name},</p>
        <p>Your teacher account has been created for the QR Attendance System.</p>
        <p><strong>Login Credentials:</strong></p>
        <p>Username: <strong>{username}</strong></p>
        <p>Temporary Password: <strong>{password}</strong></p>
        
        <div style="background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; margin: 20px 0; border-radius: 5px;">
            <h4 style="color: #856404; margin-top: 0;">⚠️ Important Security Notice</h4>
            <p style="color: #856404; margin-bottom: 0;">
                <strong>You must change your password when you first login.</strong> 
                The system will automatically redirect you to change your password for security purposes.
            </p>
        </div>
        
        <p>To get started:</p>
        <ol>
            <li>Visit the QR Attendance System</li>
            <li>Login with the credentials above</li>
            <li>You'll be prompted to create a new secure password</li>
            <li>Once changed, you can access your teacher dashboard</li>
        </ol>
        
        <p>Best regards,<br>QR Attendance System Administrator</p>
        """
        
        if send_email(email, 'Your Teacher Account Credentials', email_template):
            flash('Teacher account created and credentials sent via email')
        else:
            flash('Teacher account created but failed to send email')
        
        return redirect(url_for('admin_dashboard'))
    
    return render_template('create_teacher.html')

@app.route('/admin/delete_teacher/<teacher_id>', methods=['POST'])
@login_required
def delete_teacher(teacher_id):
    if not isinstance(current_user, Admin):
        flash('Access denied')
        return redirect(url_for('index'))
    
    teacher = Teacher.query.get_or_404(teacher_id)
    
    try:
        # Get teacher info for confirmation message
        teacher_name = teacher.full_name
        teacher_username = teacher.username
        
        # Delete the teacher (cascade will handle related records)
        db.session.delete(teacher)
        db.session.commit()
        
        flash(f'Teacher "{teacher_name}" ({teacher_username}) has been successfully deleted', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting teacher: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/teacher/dashboard')
@login_required
@teacher_password_required
def teacher_dashboard():
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    return render_template('teacher_dashboard.html', courses=courses)

@app.route('/teacher/create_course', methods=['GET', 'POST'])
@login_required
@teacher_password_required
def create_course():
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        name = request.form['name']
        code = request.form['code']
        
        if Course.query.filter_by(code=code).first():
            flash('Course code already exists')
            return render_template('create_course.html')
        
        course = Course(
            name=name,
            code=code,
            teacher_id=current_user.id
        )
        
        db.session.add(course)
        db.session.commit()
        
        registration_url = get_external_url('student_registration', course_id=course.id)
        qr_filename = f'registration_{course.id}.png'
        generate_qr_code(registration_url, qr_filename)
        
        course.registration_qr_code = qr_filename
        db.session.commit()
        
        flash('Course created successfully')
        return redirect(url_for('teacher_dashboard'))
    
    return render_template('create_course.html')

@app.route('/teacher/course/<course_id>')
@login_required
@teacher_password_required
def course_details(course_id):
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    course = Course.query.filter_by(id=course_id, teacher_id=current_user.id).first_or_404()
    students = Student.query.filter_by(course_id=course_id).all()
    sessions = AttendanceSession.query.filter_by(course_id=course_id).order_by(AttendanceSession.created_at.desc()).all()
    
    return render_template('course_details.html', course=course, students=students, sessions=sessions)

@app.route('/student/register/<course_id>', methods=['GET', 'POST'])
def student_registration(course_id):
    course = Course.query.get_or_404(course_id)
    
    if request.method == 'POST':
        name = request.form['name']
        matricule = request.form['matricule']
        sex = request.form['sex']
        
        # Check if student is already registered for THIS specific course
        existing_registration = Student.query.filter_by(matricule=matricule, course_id=course_id).first()
        if existing_registration:
            return jsonify({'error': 'You are already registered for this course'}), 400
        
        # Check if student exists in other courses to get consistent name/sex
        existing_student = Student.query.filter_by(matricule=matricule).first()
        if existing_student:
            # Use existing student's name and sex for consistency
            if existing_student.name != name or existing_student.sex != sex:
                return jsonify({
                    'error': f'Student with matricule {matricule} already exists with different details. Please use: Name: {existing_student.name}, Sex: {existing_student.sex}'
                }), 400
        
        # Create new registration for this course
        student = Student(
            name=name,
            matricule=matricule,
            sex=sex,
            course_id=course_id
        )
        
        try:
            db.session.add(student)
            db.session.commit()
            
            # Get student's total course count for response
            total_courses = Student.query.filter_by(matricule=matricule).count()
            
            return jsonify({
                'success': f'Registration successful! You are now registered for {total_courses} course(s).'
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Registration failed: {str(e)}'}), 500
    
    return render_template('student_registration.html', course=course)

@app.route('/student/profile/<matricule>')
def student_profile(matricule):
    """View all courses a student is registered for"""
    student_registrations = Student.query.filter_by(matricule=matricule).all()
    
    if not student_registrations:
        return render_template('error.html', message='No student found with this matricule')
    
    # Get the student's basic info from first registration
    student_info = student_registrations[0]
    
    # Get all courses the student is registered for
    courses = []
    for registration in student_registrations:
        course = Course.query.get(registration.course_id)
        if course:
            teacher = Teacher.query.get(course.teacher_id)
            courses.append({
                'course': course,
                'teacher': teacher,
                'registration_date': registration.registered_at
            })
    
    return render_template('student_profile.html', 
                         student_info=student_info, 
                         courses=courses,
                         total_courses=len(courses))

# Initialize database
with app.app_context():
    db.create_all()
    
    # Create default admin if it doesn't exist
    if not Admin.query.first():
        admin = Admin(
            username='admin',
            email='admin@qrattendance.com',
            password_hash=generate_password_hash('admin123')
        )
        db.session.add(admin)
        db.session.commit()
        print("Default admin created: username='admin', password='admin123'")

# Attendance Session Routes
@app.route('/teacher/create_attendance_session/<course_id>', methods=['GET', 'POST'])
@login_required
@teacher_password_required
def create_attendance_session(course_id):
    course = Course.query.get_or_404(course_id)
    
    if request.method == 'POST':
        session_name = request.form['session_name']
        latitude = request.form.get('latitude')
        longitude = request.form.get('longitude')
        
        if not latitude or not longitude:
            flash('Please capture your location before creating the session', 'error')
            return redirect(url_for('create_attendance_session', course_id=course_id))
        
        # Create attendance session with 15-minute expiry
        expires_at = datetime.utcnow() + timedelta(minutes=15)
        
        attendance_session = AttendanceSession(
            session_name=session_name,
            course_id=course_id,
            is_active=True,
            expires_at=expires_at,
            latitude=latitude,
            longitude=longitude,
            radius_meters=300  # 300 meters radius
        )
        
        db.session.add(attendance_session)
        db.session.commit()
        
        # Generate QR code for attendance
        attendance_url = get_external_url('take_attendance', session_id=attendance_session.id)
        qr_filename = f'attendance_{attendance_session.id}.png'
        generate_qr_code(attendance_url, qr_filename)
        
        attendance_session.qr_code_path = qr_filename
        db.session.commit()
        
        flash(f'Attendance session created successfully! Valid for 15 minutes until {expires_at.strftime("%H:%M")}')
        return redirect(url_for('course_details', course_id=course_id))
    
    current_date = datetime.now().strftime('%Y-%m-%d')
    return render_template('create_attendance_session.html', course=course, current_date=current_date)

@app.route('/attendance/<session_id>', methods=['GET', 'POST'])
def take_attendance(session_id):
    session = AttendanceSession.query.get_or_404(session_id)
    
    # Check if session is expired
    if session.is_expired():
        return render_template('error.html', 
                             message=f'This attendance session has expired. Sessions are only valid for 15 minutes.')
    
    if not session.is_active:
        return render_template('error.html', message='This attendance session is no longer active')
    
    course = Course.query.get(session.course_id)
    students = Student.query.filter_by(course_id=course.id).all()
    
    if request.method == 'POST':
        selected_matricule = request.form['selected_matricule']
        
        # Find student by matricule in this course
        student = Student.query.filter_by(matricule=selected_matricule, course_id=course.id).first()
        if not student:
            return jsonify({'error': 'Invalid matricule selection'}), 400
        
        # Check if student already marked attendance
        existing_attendance = Attendance.query.filter_by(
            student_id=student.id,
            session_id=session_id
        ).first()
        
        if existing_attendance:
            return jsonify({'error': 'Attendance already marked for this session'}), 400
        
        # Get student's location
        try:
            student_latitude = float(request.form.get('latitude', 0))
            student_longitude = float(request.form.get('longitude', 0))
            
            # Verify location (within 300m radius)
            if session.latitude and session.longitude:
                distance = calculate_distance(
                    float(session.latitude), float(session.longitude),
                    student_latitude, student_longitude
                )
                
                if distance > session.radius_meters:
                    return jsonify({
                        'error': f'You are {distance:.0f}m away from the class location. Please move closer (within {session.radius_meters}m)'
                    }), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid location data'}), 400
        
        # Mark attendance
        attendance = Attendance(
            student_id=student.id,
            session_id=session_id,
            latitude=student_latitude,
            longitude=student_longitude
        )
        
        db.session.add(attendance)
        db.session.commit()
        
        return jsonify({
            'success': f'Attendance marked successfully!',
            'student_name': student.name,
            'student_matricule': student.matricule,
            'student_sex': student.sex,
            'time_remaining': session.minutes_remaining()
        })
    
    return render_template('take_attendance.html', session=session, course=course, students=students)

@app.route('/teacher/course_analytics/<course_id>')
@login_required
@teacher_password_required
def course_analytics(course_id):
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    course = Course.query.filter_by(id=course_id, teacher_id=current_user.id).first_or_404()
    students = Student.query.filter_by(course_id=course_id).all()
    sessions = AttendanceSession.query.filter_by(course_id=course_id).all()
    
    # Calculate statistics
    total_students = len(students)
    total_sessions = len(sessions)
    
    # Gender distribution
    male_students = len([s for s in students if s.sex == 'Male'])
    female_students = len([s for s in students if s.sex == 'Female'])
    
    # Attendance statistics
    attendance_stats = []
    for student in students:
        student_attendances = Attendance.query.join(AttendanceSession).filter(
            Attendance.student_id == student.id,
            AttendanceSession.course_id == course_id
        ).count()
        
        attendance_rate = (student_attendances / total_sessions * 100) if total_sessions > 0 else 0
        attendance_stats.append({
            'student': student,
            'attendance_count': student_attendances,
            'attendance_rate': attendance_rate
        })
    
    return render_template('course_analytics.html', 
                         course=course, 
                         students=students, 
                         sessions=sessions,
                         total_students=total_students,
                         total_sessions=total_sessions,
                         male_students=male_students,
                         female_students=female_students,
                         attendance_stats=attendance_stats)

@app.route('/teacher/end_session/<session_id>')
@login_required
@teacher_password_required
def end_attendance_session(session_id):
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    session = AttendanceSession.query.get_or_404(session_id)
    course = Course.query.get(session.course_id)
    
    # Verify teacher owns this course
    if course.teacher_id != current_user.id:
        flash('Access denied')
        return redirect(url_for('teacher_dashboard'))
    
    # End the session
    session.is_active = False
    db.session.commit()
    
    flash(f'Attendance session "{session.session_name}" has been ended', 'success')
    return redirect(url_for('course_details', course_id=course.id))

@app.route('/admin/regenerate_qr_codes')
@login_required
def regenerate_qr_codes():
    """Regenerate all QR codes with the current server IP"""
    if not isinstance(current_user, Admin):
        flash('Access denied')
        return redirect(url_for('index'))
    
    try:
        # Regenerate registration QR codes for all courses
        courses = Course.query.all()
        courses_updated = 0
        
        for course in courses:
            if course.registration_qr_code:
                # Generate new registration QR code
                registration_url = get_external_url('student_registration', course_id=course.id)
                qr_filename = f'registration_{course.id}.png'
                generate_qr_code(registration_url, qr_filename)
                
                course.registration_qr_code = qr_filename
                courses_updated += 1
        
        # Regenerate attendance QR codes for all active sessions
        sessions = AttendanceSession.query.filter_by(is_active=True).all()
        sessions_updated = 0
        
        for session in sessions:
            if session.qr_code_path:
                # Generate new attendance QR code
                attendance_url = get_external_url('take_attendance', session_id=session.id)
                qr_filename = f'attendance_{session.id}.png'
                generate_qr_code(attendance_url, qr_filename)
                
                session.qr_code_path = qr_filename
                sessions_updated += 1
        
        db.session.commit()
        
        flash(f'Successfully regenerated QR codes for {courses_updated} courses and {sessions_updated} active attendance sessions with current server IP: {app.config["SERVER_IP"]}')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error regenerating QR codes: {str(e)}')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/teacher/students')
@login_required
@teacher_password_required
def teacher_students():
    """View all students across all teacher's courses"""
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    # Get all courses taught by this teacher
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    
    # Get all students from all courses
    all_students = []
    unique_matricules = set()
    
    for course in courses:
        course_students = Student.query.filter_by(course_id=course.id).all()
        for student in course_students:
            all_students.append({
                'student': student,
                'course': course
            })
            unique_matricules.add(student.matricule)
    
    # Group students by matricule for summary
    student_summary = []
    for matricule in unique_matricules:
        student_courses = [item for item in all_students if item['student'].matricule == matricule]
        student_info = student_courses[0]['student']  # Get basic info
        student_summary.append({
            'matricule': matricule,
            'name': student_info.name,
            'sex': student_info.sex,
            'total_courses': len(student_courses),
            'courses': [item['course'].name for item in student_courses]
        })
    
    return render_template('teacher_students.html', 
                         all_students=all_students,
                         student_summary=student_summary,
                         courses=courses,
                         total_unique_students=len(unique_matricules))

@app.route('/mobile-qr-test')
def mobile_qr_test():
    """Generate test QR codes for mobile testing"""
    # Generate a test QR code for mobile scanning
    test_url = f"http://{app.config['SERVER_IP']}:{app.config['SERVER_PORT']}/mobile-test"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(test_url)
    qr.make(fit=True)
    
    # Save the QR code
    import io
    import base64
    img = qr.make_image(fill_color="black", back_color="white")
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_str = base64.b64encode(img_buffer.getvalue()).decode()
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>QR Test - Mobile Scanner</title>
        <style>
            body {{ 
                font-family: Arial, sans-serif; 
                padding: 20px; 
                text-align: center; 
                background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
                color: white;
                min-height: 100vh;
                margin: 0;
            }}
            .container {{
                max-width: 500px;
                margin: 0 auto;
                background: white;
                color: #333;
                padding: 30px;
                border-radius: 15px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.3);
            }}
            .qr-code {{
                background: white;
                padding: 20px;
                border-radius: 10px;
                margin: 20px 0;
                display: inline-block;
                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            }}
            .instructions {{
                background: #f8f9fa;
                padding: 20px;
                border-radius: 10px;
                margin: 20px 0;
                text-align: left;
                border-left: 4px solid #28a745;
            }}
            .btn {{
                background: #28a745;
                color: white;
                padding: 15px 30px;
                border: none;
                border-radius: 25px;
                text-decoration: none;
                display: inline-block;
                margin: 10px;
                font-size: 16px;
                transition: all 0.3s;
            }}
            .btn:hover {{
                background: #1e7e34;
                transform: translateY(-2px);
            }}
            h1 {{ margin-top: 0; color: #28a745; }}
            @media (max-width: 480px) {{
                body {{ padding: 10px; }}
                .container {{ padding: 20px; }}
                .qr-code img {{ max-width: 100%; height: auto; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📱 QR Code Mobile Test</h1>
            <p><strong>Test QR Code for Mobile Scanning</strong></p>
            
            <div class="qr-code">
                <img src="data:image/png;base64,{img_str}" alt="Test QR Code" style="max-width: 300px;">
            </div>
            
            <div class="instructions">
                <h3>📋 How to Test:</h3>
                <ol>
                    <li><strong>Open QR Scanner:</strong> Use your phone's camera or QR scanner app</li>
                    <li><strong>Scan QR Code:</strong> Point your camera at the QR code above</li>
                    <li><strong>Tap the Link:</strong> Your phone should open the mobile test page</li>
                    <li><strong>Verify Access:</strong> You should see a success message</li>
                </ol>
            </div>
            
            <div class="instructions">
                <h3>✅ What This Tests:</h3>
                <ul>
                    <li>QR code generation with correct network IP</li>
                    <li>Mobile device network connectivity</li>
                    <li>Mobile browser compatibility</li>
                    <li>Touch interface responsiveness</li>
                </ul>
            </div>
            
            <p><strong>Target URL:</strong><br>
            <code>{test_url}</code></p>
            
            <a href="/" class="btn">🏠 Back to Login</a>
            <a href="/mobile-test" class="btn">📱 Direct Mobile Test</a>
        </div>
    </body>
    </html>
    """

@app.route('/verify-qr-codes')
def verify_qr_codes():
    """Debug route to verify all QR codes are using correct IP"""
    courses = Course.query.all()
    sessions = AttendanceSession.query.all()
    
    course_qr_info = []
    for course in courses:
        registration_url = get_external_url('student_registration', course_id=course.id)
        course_qr_info.append({
            'course_name': course.name,
            'course_id': course.id,
            'registration_url': registration_url,
            'qr_file': course.registration_qr_code
        })
    
    session_qr_info = []
    for session in sessions:
        attendance_url = get_external_url('take_attendance', session_id=session.id)
        session_qr_info.append({
            'session_name': session.session_name,
            'session_id': session.id,
            'attendance_url': attendance_url,
            'qr_file': session.qr_code_path,
            'is_active': session.is_active
        })
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>QR Code Verification</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }}
            .container {{ max-width: 1200px; margin: 0 auto; }}
            .status {{ padding: 15px; border-radius: 5px; margin: 20px 0; }}
            .success {{ background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }}
            .info {{ background: #cce7ff; color: #004085; border: 1px solid #b3d7ff; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
            th {{ background-color: #f8f9fa; font-weight: bold; }}
            .url {{ font-family: monospace; font-size: 12px; color: #007bff; }}
            .btn {{ background: #007bff; color: white; padding: 8px 16px; text-decoration: none; border-radius: 4px; margin: 2px; }}
            .btn:hover {{ background: #0056b3; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🔍 QR Code Verification Dashboard</h1>
            
            <div class="status success">
                <strong>✅ Server Configuration</strong><br>
                IP Address: <code>{app.config['SERVER_IP']}</code><br>
                Port: <code>{app.config['SERVER_PORT']}</code><br>
                Base URL: <code>http://{app.config['SERVER_IP']}:{app.config['SERVER_PORT']}</code>
            </div>
            
            <h2>📚 Course Registration QR Codes</h2>
            <table>
                <thead>
                    <tr>
                        <th>Course Name</th>
                        <th>Course ID</th>
                        <th>Registration URL</th>
                        <th>QR File</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([f'''
                    <tr>
                        <td><strong>{course['course_name']}</strong></td>
                        <td><code>{course['course_id'][:8]}...</code></td>
                        <td class="url">{course['registration_url']}</td>
                        <td>{course['qr_file'] or 'No QR code'}</td>
                        <td>
                            <a href="{course['registration_url']}" class="btn" target="_blank">Test URL</a>
                        </td>
                    </tr>
                    ''' for course in course_qr_info])}
                </tbody>
            </table>
            
            <h2>📅 Attendance Session QR Codes</h2>
            <table>
                <thead>
                    <tr>
                        <th>Session Name</th>
                        <th>Session ID</th>
                        <th>Attendance URL</th>
                        <th>QR File</th>
                        <th>Status</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join([f'''
                    <tr>
                        <td><strong>{session['session_name']}</strong></td>
                        <td><code>{session['session_id'][:8]}...</code></td>
                        <td class="url">{session['attendance_url']}</td>
                        <td>{session['qr_file'] or 'No QR code'}</td>
                        <td>{'🟢 Active' if session['is_active'] else '🔴 Inactive'}</td>
                        <td>
                            <a href="{session['attendance_url']}" class="btn" target="_blank">Test URL</a>
                        </td>
                    </tr>
                    ''' for session in session_qr_info])}
                </tbody>
            </table>
            
            <div class="status info">
                <strong>📱 Mobile Test Links:</strong><br>
                <a href="http://{app.config['SERVER_IP']}:{app.config['SERVER_PORT']}/mobile-test" class="btn">Mobile Test</a>
                <a href="http://{app.config['SERVER_IP']}:{app.config['SERVER_PORT']}/mobile-qr-test" class="btn">QR Test</a>
                <a href="/" class="btn">Back to System</a>
            </div>
            
            <p><small>Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</small></p>
        </div>
    </body>
    </html>
    """

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        # Verify current password
        if not check_password_hash(current_user.password_hash, current_password):
            flash('Current password is incorrect', 'error')
            return render_template('change_password.html')
        
        # Validate new password
        if len(new_password) < 8:
            flash('New password must be at least 8 characters long', 'error')
            return render_template('change_password.html')
        
        if new_password != confirm_password:
            flash('Password confirmation does not match', 'error')
            return render_template('change_password.html')
        
        # Update password
        current_user.password_hash = generate_password_hash(new_password)
        current_user.must_change_password = False
        db.session.commit()
        
        flash('Password changed successfully!', 'success')
        return redirect(url_for('teacher_dashboard'))
    
    return render_template('change_password.html')

@app.route('/teacher/export_attendance/<session_id>/<format>')
@login_required
@teacher_password_required
def export_attendance(session_id, format):
    if not isinstance(current_user, Teacher):
        flash('Access denied')
        return redirect(url_for('index'))
    
    session = AttendanceSession.query.get_or_404(session_id)
    course = Course.query.get(session.course_id)
    
    # Verify teacher owns this course
    if course.teacher_id != current_user.id:
        flash('Access denied')
        return redirect(url_for('teacher_dashboard'))
    
    # Get attendance records and all students
    attendances = Attendance.query.filter_by(session_id=session_id).all()
    students = Student.query.filter_by(course_id=course.id).all()
    
    # Prepare data
    attendance_data = []
    for student in students:
        attendance = next((a for a in attendances if a.student_id == student.id), None)
        attendance_data.append({
            'Name': student.name,
            'Matricule': student.matricule,
            'Sex': student.sex,
            'Status': 'Present' if attendance else 'Absent',
            'Time Marked': attendance.marked_at.strftime('%H:%M:%S') if attendance else 'N/A',
            'Date': attendance.marked_at.strftime('%Y-%m-%d') if attendance else 'N/A'
        })
    
    if format == 'excel':
        return export_to_excel(attendance_data, session, course)
    elif format == 'pdf':
        return export_to_pdf(attendance_data, session, course)
    else:
        flash('Invalid export format')
        return redirect(url_for('course_details', course_id=course.id))

def export_to_excel(attendance_data, session, course):
    """Export attendance to Excel file"""
    try:
        # Create DataFrame
        df = pd.DataFrame(attendance_data)
        
        # Create filename
        filename = f"attendance_{course.name}_{session.session_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('static', 'exports', filename)
        
        # Ensure exports directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance', index=False)
            
            # Get worksheet for formatting
            worksheet = writer.sheets['Attendance']
            
            # Add header with course info
            worksheet.insert_rows(1, 3)
            worksheet['A1'] = f"Course: {course.name} ({course.code})"
            worksheet['A2'] = f"Session: {session.session_name}"
            worksheet['A3'] = f"Date: {session.created_at.strftime('%Y-%m-%d')}"
            worksheet['A4'] = f"Teacher: {course.teacher.full_name}"
            
            # Style the header
            try:
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for row in range(1, 5):
                    for col in range(1, 7):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = header_font
                        if row <= 4:
                            cell.fill = header_fill
            except ImportError:
                # If openpyxl is not available, continue without styling
                print("Warning: openpyxl styling not available")
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'Error creating Excel file: {str(e)}')
        return redirect(url_for('course_details', course_id=course.id))

def export_to_pdf(attendance_data, session, course):
    """Export attendance to PDF file"""
    try:
        # Create filename
        filename = f"attendance_{course.name}_{session.session_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join('static', 'exports', filename)
        
        # Ensure exports directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Create PDF
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(f"<b>Attendance Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Course Information
        course_info = f"""
        <b>Course:</b> {course.name} ({course.code})<br/>
        <b>Session:</b> {session.session_name}<br/>
        <b>Date:</b> {session.created_at.strftime('%Y-%m-%d')}<br/>
        <b>Teacher:</b> {course.teacher.full_name}<br/>
        <b>Total Present:</b> {len(attendance_data)}<br/>
        <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        course_para = Paragraph(course_info, styles['Normal'])
        elements.append(course_para)
        elements.append(Spacer(1, 24))
        
        # Attendance Table
        if attendance_data:
            # Table headers
            table_data = [['#', 'Name', 'Matricule', 'Sex', 'Time Marked']]
            
            # Table rows
            for i, record in enumerate(attendance_data, 1):
                table_data.append([
                    str(i),
                    record['Name'],
                    record['Matricule'],
                    record['Sex'],
                    record['Time Marked']
                ])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        else:
            no_attendance = Paragraph("<b>No attendance records found for this session.</b>", styles['Normal'])
            elements.append(no_attendance)
        
        # Build PDF
        doc.build(elements)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'Error creating PDF file: {str(e)}')
        return redirect(url_for('course_details', course_id=course.id))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("\n" + "="*70)
        print("🚀 QR ATTENDANCE SYSTEM - NETWORK & MOBILE READY")
        print("="*70)
        print(f"💻 Local Access:      http://localhost:5000")
        print(f"🌐 Network Access:    http://{app.config['SERVER_IP']}:5000")
        print(f"📱 Mobile Access:     http://{app.config['SERVER_IP']}:5000")
        print(f"🔗 Teacher Access:    http://{app.config['SERVER_IP']}:5000")
        print("="*70)
        print("📋 QUICK START GUIDE:")
        print("1. Admin: Login with code 23456")
        print("2. Create teacher accounts (credentials sent via email)")
        print("3. Teachers: Login and create courses")
        print("4. Generate QR codes for student registration/attendance")
        print("5. Students: Scan QR codes on mobile phones!")
        print("="*70)
        print("📱 MOBILE TESTING:")
        print(f"   Test URL: http://{app.config['SERVER_IP']}:5000/mobile-test")
        print("   • Students can scan QR codes from their phones")
        print("   • Location-based attendance verification")
        print("   • Mobile-optimized registration forms")
        print("="*70)
        print("🔧 NETWORK SETUP:")
        print(f"   • Server IP: {app.config['SERVER_IP']}")
        print("   • All devices must be on the same WiFi network")
        print("   • QR codes automatically use network IP")
        print("   • Teachers can access from their own devices")
        print("="*70)
    
    # Run with network access enabled
    app.run(
        host='0.0.0.0',  # Listen on all network interfaces
        port=5000,
        debug=True,
        threaded=True
    ) 